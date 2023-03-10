# Importing necessary libraries
import csv
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# Reading HTML file and defining paths for CSV and Excel files
file = pd.read_html("./Test Report_2021-08-18_12-45-00.html")
path = "./your_csv_name.csv"
xlpath = 'name.xlsx'

# Function to write data from HTML to CSV and convert it to Excel format
def write_html_csv():
    for index, data in enumerate(file):
        # Check for index value and print data
        if index:
            data.to_csv("./your_csv_name.csv", mode='a+', header=True)

    # Creating an instance of Workbook and creating a new sheet
    wb = Workbook()
    ws = wb.active

    # Reading CSV file and writing data to Excel
    with open(path, 'r') as f:
        for row in csv.reader(f):
            ws.append(row)
    
    # Saving the Excel file
    wb.save(xlpath)

# Function to modify the Excel sheet by adding bold font to certain cell values
def modify_excel():
    # Opening the Excel file
    wb_obj = openpyxl.load_workbook(xlpath)
    sheet_obj = wb_obj.active

    # Getting the number of rows and columns in the sheet
    rows = sheet_obj.max_row
    cols = sheet_obj.max_column

    # Looping through each cell and checking for certain values to apply font style
    for i in range(1, rows + 1):
        for j in range(1, cols + 1):
            if ("Test_Cases" in str(sheet_obj.cell(i, j).value)) or ("Status" in str(sheet_obj.cell(i, j).value)):
                x = sheet_obj.cell(i, j).coordinate
                y = sheet_obj.cell(i, j).row
                sheet_obj[x].font = Font(bold=True)

    # Saving the modified Excel file
    wb_obj.save(xlpath)

# Running the functions and printing messages to indicate completion of tasks
print("Starting task one")
write_html_csv()
print("Task one over")
print("Starting task two")
modify_excel()
print("Task two over")
