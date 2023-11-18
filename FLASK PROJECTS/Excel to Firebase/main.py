import firebase_admin
from flask import Flask,render_template,request,redirect
from flask.helpers import url_for
from openpyxl import load_workbook
import pandas as pd
import numpy as np
# from firebase_admin import db
# from firebase_admin import credentials
from datetime import date
import datetime
from firebase import Firebase
import os

firebaseConfig = {
  "apiKey": "your apikey",
  "authDomain": "",
  "databaseURL": "database url",
  "projectId": "your project id",
  "storageBucket": "your storage bucket id",
  "messagingSenderId": "your sender id",
  "appId": "your appId",
  "measurementId": "your measurement id"
};
firebase = Firebase(firebaseConfig)
db = firebase.database()
list_of_keys=[]

app = Flask(__name__)

@app.route("/",methods = ['GET','POST'])
def start():
    if request.method == 'POST':
        global file,s_ds1
        file = request.files['file']
        #file.save("random.xlsx")
        file.save("static/police_record_today.xlsx")
        global wb
        wb = load_workbook(file)  # Work Book
          # Work Sheet MCRC.RM.COLL 
        sheets=wb.sheetnames

        return render_template("sheetname.html",sheets = sheets)
    return render_template("index.html")

@app.route("/sheet-selection",methods = ['GET','POST'])
def sheet_selection():
    if request.method == 'POST':
        sheet_name_user = request.form.get("sheet_selected")
        print(sheet_name_user)
        today = date.today()
        ds1 = pd.read_excel("static/police_record_today.xlsx", sheet_name_user)
        ds2=pd.read_excel("static/police_record_yesterday.xlsx",sheet_name_user)
        #appending both data
        merged = ds1.append(ds2)
        #dropping duplicates
        merged = merged.drop_duplicates(keep=False).sort_index()
        #converting to string 
        data=merged.values.tolist()
        for i in range(len(data)):
            if(str(data[i][10])!="nan"):
                fd=datetime.datetime.strptime(str(data[i][10]), '%d.%m.%Y').strftime('%Y.%m.%d').replace(".","")
                key=fd+str(data[i][1])+str(data[i][2])+str(data[i][3])+str(data[i][4])+str(data[i][6])+str(data[i][7])+str(data[i][8]);
                print(key)
                push_key=key.replace(".","").replace(" ","")
                if(str(data[i][1])=="nan" and str(data[i][2])=="nan"):
                    continue
                ct=str(data[i][3])
                cn=str(data[i][4]).replace(".0","")
                name=str(data[i][5])
                ca_yr=str(data[i][6]).replace(".0","")
                crn=str(data[i][7]).replace(".0","")
                cr_yr=str(data[i][8]).replace(".0","")
                d_r=str(data[i][9])
                rm_date=str(data[i][10])
                before=str(data[i][11])

                if(str(data[i][3])=="nan"):
                    ct="None"
                if(str(data[i][4])=="nan"):
                    cn="None"
                if(str(data[i][5])=="nan"):
                    name="None"
                if(str(data[i][6])=="nan"):
                    ca_yr="None"
                if(str(data[i][7])=="nan"):
                    crn="None"
                if(str(data[i][8])=="nan"):
                    cr_yr="None"
                if(str(data[i][9])=="nan"):
                    d_r="None"
                if(str(data[i][10])=="nan"):
                    rm_date="None"
                if(str(data[i][11])=="nan"):
                    before="None"
                diction = {
                            'A':"",
                            'B':str(data[i][1]),
                            'C':str(data[i][2]),
                            'D':ct,
                            'E':cn,
                            'F':name,
                            'G':ca_yr,
                            'H':crn,
                            'I':cr_yr,
                            'J':d_r,
                            'K':rm_date,
                            'L':before,
                            'date':str(today),
                            'pushkey':push_key,
                            'type':sheet_name_user.replace(".","_").strip()
                        }
                db.child('data').child(push_key).set(diction)
        os.remove("static/police_record_yesterday.xlsx")
        os.rename("static/police_record_today.xlsx","static/police_record_yesterday.xlsx")
        #s_ds1.to_excel("static/police_record_yesterday.xlsx",index=False)#save today file as tomorrow
        

        return redirect(url_for('start'))



if __name__ == '__main__':
    app.run(debug = True)

